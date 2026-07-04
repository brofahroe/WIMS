from __future__ import annotations

import json
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "WIMS V1_GCI-EJ-EMR-MALANG.xlsm"
OUT_DIR = ROOT / ".codex_tmp"

NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
NS_PKG_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"


def json_safe(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def cell_preview(ws, max_rows=24, max_cols=18):
    rows = []
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(max_rows, ws.max_row or 1),
        min_col=1,
        max_col=min(max_cols, ws.max_column or 1),
        values_only=False,
    ):
        out = []
        for cell in row:
            value = cell.value
            if value is None:
                out.append("")
            elif isinstance(value, str):
                out.append(value[:120])
            else:
                out.append(json_safe(value))
        rows.append(out)
    return rows


def non_empty_bounds(ws):
    min_row = min_col = None
    max_row = max_col = None
    count = 0
    formulas = 0
    strings = 0
    numbers = 0
    bools = 0
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            count += 1
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_col = cell.column if max_col is None else max(max_col, cell.column)
            if isinstance(v, str) and v.startswith("="):
                formulas += 1
            elif isinstance(v, str):
                strings += 1
            elif isinstance(v, bool):
                bools += 1
            elif isinstance(v, (int, float)):
                numbers += 1
    return {
        "non_empty_cells": count,
        "formula_cells": formulas,
        "string_cells": strings,
        "number_cells": numbers,
        "bool_cells": bools,
        "bounds": [min_row, min_col, max_row, max_col],
    }


def formula_samples(ws, limit=30):
    samples = []
    functions = Counter()
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                samples.append({"cell": cell.coordinate, "formula": v[:250]})
                for fn in re.findall(r"([A-Z][A-Z0-9_\.]*)\s*\(", v.upper()):
                    functions[fn] += 1
    return samples[:limit], functions


def workbook_xml_summary():
    summary = {
        "zip_entries": [],
        "workbook_sheets": [],
        "defined_names": [],
        "rels": [],
        "drawing_entries": [],
        "control_entries": [],
        "vba_project_present": False,
    }
    with zipfile.ZipFile(WORKBOOK) as z:
        for info in z.infolist():
            summary["zip_entries"].append(
                {"name": info.filename, "size": info.file_size, "compressed": info.compress_size}
            )
            lower = info.filename.lower()
            if "vba" in lower:
                summary["vba_project_present"] = True
            if lower.startswith("xl/drawings/") or lower.startswith("xl/media/"):
                summary["drawing_entries"].append(info.filename)
            if "ctrlprops" in lower or "controls" in lower or "active" in lower:
                summary["control_entries"].append(info.filename)

        wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
        for sheet in wb_xml.findall(f"{NS_MAIN}sheets/{NS_MAIN}sheet"):
            summary["workbook_sheets"].append(
                {
                    "name": sheet.attrib.get("name"),
                    "sheetId": sheet.attrib.get("sheetId"),
                    "state": sheet.attrib.get("state", "visible"),
                    "rId": sheet.attrib.get(f"{NS_REL}id"),
                }
            )
        for dn in wb_xml.findall(f"{NS_MAIN}definedNames/{NS_MAIN}definedName"):
            text = dn.text or ""
            summary["defined_names"].append(
                {
                    "name": dn.attrib.get("name"),
                    "localSheetId": dn.attrib.get("localSheetId"),
                    "text": text[:300],
                }
            )
        rels_path = "xl/_rels/workbook.xml.rels"
        if rels_path in z.namelist():
            rels = ET.fromstring(z.read(rels_path))
            for rel in rels.findall(f"{NS_PKG_REL}Relationship"):
                summary["rels"].append(rel.attrib)
    return summary


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False, keep_vba=True)
    data_wb = openpyxl.load_workbook(WORKBOOK, data_only=True, keep_vba=False)

    out = {
        "workbook": WORKBOOK.name,
        "sheet_count": len(wb.sheetnames),
        "sheets": [],
        "xml": workbook_xml_summary(),
    }

    all_functions = Counter()
    for ws in wb.worksheets:
        samples, functions = formula_samples(ws)
        all_functions.update(functions)
        data_ws = data_wb[ws.title]
        merged = [str(rng) for rng in ws.merged_cells.ranges]
        validations = []
        try:
            for dv in ws.data_validations.dataValidation:
                validations.append({"type": dv.type, "sqref": str(dv.sqref), "formula1": dv.formula1})
        except Exception:
            pass
        tables = []
        try:
            for name, table in ws.tables.items():
                tables.append({"name": name, "ref": getattr(table, "ref", None)})
        except Exception:
            pass
        charts = []
        try:
            charts = [type(ch).__name__ for ch in ws._charts]
        except Exception:
            pass
        images = []
        try:
            images = [type(img).__name__ for img in ws._images]
        except Exception:
            pass
        out["sheets"].append(
            {
                "title": ws.title,
                "state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "non_empty": non_empty_bounds(ws),
                "merged_ranges": merged[:80],
                "merged_count": len(merged),
                "tables": tables,
                "data_validations": validations[:40],
                "charts": charts,
                "images": images,
                "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                "formula_samples": samples,
                "top_formula_functions": functions.most_common(20),
                "preview_formulas": cell_preview(ws),
                "preview_values": cell_preview(data_ws),
            }
        )
    out["top_formula_functions"] = all_functions.most_common(50)

    OUT_DIR.mkdir(exist_ok=True)
    json_path = OUT_DIR / "workbook_audit.json"
    json_path.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Workbook: {out['workbook']}")
    print(f"Sheets: {out['sheet_count']}")
    for sheet in out["sheets"]:
        ne = sheet["non_empty"]
        print(
            f"- {sheet['title']} [{sheet['state']}]: "
            f"{ne['non_empty_cells']} cells, {ne['formula_cells']} formulas, "
            f"bounds={ne['bounds']}, merged={sheet['merged_count']}, "
            f"tables={len(sheet['tables'])}, validations={len(sheet['data_validations'])}, "
            f"charts={len(sheet['charts'])}, images={len(sheet['images'])}"
        )
    print("Top formula functions:", out["top_formula_functions"][:20])
    print(f"Audit JSON: {json_path}")


if __name__ == "__main__":
    sys.exit(main())
