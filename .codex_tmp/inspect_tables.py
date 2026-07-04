from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "WIMS V1_GCI-EJ-EMR-MALANG.xlsm"
OUT = ROOT / ".codex_tmp" / "tables_summary.json"


def value_of(cell):
    value = cell.value
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def table_matrix(ws, ref, max_rows=12):
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    rows = []
    for r in range(min_row, min(max_row, min_row + max_rows - 1) + 1):
        rows.append([value_of(ws.cell(r, c)) for c in range(min_col, max_col + 1)])
    return rows


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, keep_vba=False)
    wb_formula = openpyxl.load_workbook(WORKBOOK, data_only=False, keep_vba=True)
    summary = {}
    for ws in wb.worksheets:
        formula_ws = wb_formula[ws.title]
        sheet_tables = []
        for table in formula_ws.tables.values():
            ref = table.ref
            min_col, min_row, max_col, max_row = range_boundaries(ref)
            headers = [value_of(formula_ws.cell(min_row, c)) for c in range(min_col, max_col + 1)]
            sheet_tables.append(
                {
                    "name": table.name,
                    "ref": ref,
                    "row_count": max_row - min_row,
                    "col_count": max_col - min_col + 1,
                    "headers": headers,
                    "sample_values": table_matrix(ws, ref, max_rows=8),
                    "sample_formulas": table_matrix(formula_ws, ref, max_rows=8),
                }
            )
        summary[ws.title] = {
            "state": ws.sheet_state,
            "tables": sheet_tables,
        }
    OUT.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    for sheet, info in summary.items():
        if info["tables"]:
            print(f"## {sheet}")
            for table in info["tables"]:
                print(f"- {table['name']} {table['ref']} rows={table['row_count']} cols={table['col_count']}")
                print("  headers:", " | ".join("" if h is None else str(h) for h in table["headers"]))
                for row in table["sample_values"][:4]:
                    print("  row:", row)
            print()
    print(f"Saved {OUT}")


if __name__ == "__main__":
    main()
