from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "WIMS V1_GCI-EJ-EMR-MALANG.xlsm"
OUT = ROOT / "src" / "data" / "seedData.json"


def clean_key(value: str) -> str:
    return " ".join(str(value).replace("_", " ").replace("/", " ").replace(".", "").split()).lower()


def value_of(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float, bool, str)):
        if isinstance(value, str):
            return value.replace("\xa0", " ").strip()
        return value
    return str(value)


def row_has_data(row: dict[str, Any]) -> bool:
    return any(v not in (None, "") for v in row.values())


def read_table(wb, sheet_name: str, table_name: str) -> list[dict[str, Any]]:
    ws = wb[sheet_name]
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    headers = [value_of(ws.cell(min_row, col).value) or f"Column {col}" for col in range(min_col, max_col + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(min_row + 1, max_row + 1):
        record = {
            str(headers[col_idx - min_col]): value_of(ws.cell(row_idx, col_idx).value)
            for col_idx in range(min_col, max_col + 1)
        }
        if row_has_data(record):
            rows.append(record)
    return rows


def pick(row: dict[str, Any], *names: str) -> Any:
    lookup = {clean_key(k): v for k, v in row.items()}
    for name in names:
        if clean_key(name) in lookup:
            return lookup[clean_key(name)]
    return None


def unique(values: list[Any]) -> list[Any]:
    seen = set()
    out = []
    for value in values:
        if value in (None, ""):
            continue
        key = str(value).strip()
        if key and key not in seen:
            seen.add(key)
            out.append(value)
    return out


def as_number(value: Any) -> float:
    if value in (None, ""):
        return 0
    try:
        return float(value)
    except Exception:
        return 0


def map_material(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "rowId": pick(row, "Row_ID"),
        "typeMaterial": pick(row, "Type Material"),
        "sourceMaterial": pick(row, "Source Material"),
        "materialCode": pick(row, "Material Code"),
        "materialName": pick(row, "Material Name"),
        "unit": pick(row, "Unit"),
        "inbound": as_number(pick(row, "INBOUND")),
        "outbound": as_number(pick(row, "OUTBOUND")),
        "transferIn": as_number(pick(row, "Transfer IN")),
        "transferOut": as_number(pick(row, "Transfer Out")),
        "borrowIn": as_number(pick(row, "Borrow IN")),
        "borrowOut": as_number(pick(row, "Borrow Out")),
        "stockWh": as_number(pick(row, "Stock WH")),
        "leftoversStock": as_number(pick(row, "Leftovers Stock")),
        "addRemark": pick(row, "ADD REMARK"),
    }


def map_do(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "siteId": pick(row, "SiteID"),
        "siteName": pick(row, "SiteName"),
        "subcon": pick(row, "Subcon"),
        "region": pick(row, "REGION"),
        "city": pick(row, "City"),
        "dropCity": pick(row, "Drop to CITY"),
        "doNumber": pick(row, "DO Number"),
        "dnNumber": pick(row, "DN Number"),
        "materialPickUpdate": pick(row, "Material Pick-update"),
        "materialName": pick(row, "Material Name"),
        "qty": as_number(pick(row, "Qty")),
    }


def map_site(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "no": pick(row, "NO"),
        "region": pick(row, "Region"),
        "city": pick(row, "City"),
        "siteId": pick(row, "Site ID"),
        "siteName": pick(row, "Sitename"),
        "address": pick(row, "Address / Grouping"),
        "team": pick(row, "Team"),
        "permit": pick(row, "PERMIT"),
        "snd": pick(row, "SND"),
        "donation": pick(row, "DONATION"),
        "implementation": pick(row, "IMPLEMENTATION"),
        "atp": pick(row, "ATP"),
        "acceptance": pick(row, "ACCEPTANCE"),
        "finalMilestone": pick(row, "FINAL MILESTONE"),
        "materialRequest": pick(row, "Material Request"),
        "milestoneByZte": pick(row, "Milestone by ZTE"),
        "projectName": pick(row, "Project Name"),
        "statusCity": pick(row, "Status City"),
        "materials": {
            "Pole7M4": as_number(pick(row, 'Pole7M4"')),
            "Pole7M3": as_number(pick(row, 'Pole7M3"')),
            "Pole9M": as_number(pick(row, "Pole9M")),
            "ADSS24C": as_number(pick(row, "ADSS24C")),
            "ADSS48C": as_number(pick(row, "ADSS48C")),
            "ADSS96C": as_number(pick(row, "ADSS96C")),
            "FDT48": as_number(pick(row, "FDT-48")),
            "FDT72": as_number(pick(row, "FDT-72")),
            "FAT16": as_number(pick(row, "FAT-16")),
        },
    }


def map_transaction(row: dict[str, Any], source: str) -> dict[str, Any]:
    return {
        "id": f"{source}-{pick(row, 'Row_ID') or pick(row, 'Tag_ID') or len(str(row))}",
        "source": source,
        "rowId": pick(row, "Row_ID"),
        "tagId": pick(row, "Tag_ID"),
        "lineId": pick(row, "Line_ID", "Line ID"),
        "taggingType": "LEFTOVERS" if source == "leftovers" else "LOGFILE",
        "transactionType": pick(row, "Transaction Type", "Transaction  Type"),
        "notaNo": pick(row, "Nota No", "Nota No."),
        "whGci": pick(row, "WH GCI"),
        "picWarehouse": pick(row, "PIC Warehouse"),
        "date": pick(row, "Date"),
        "time": pick(row, "Time"),
        "sourceDestination": pick(row, "Material Source / Destination"),
        "typeMaterial": pick(row, "Type Material"),
        "materialName": pick(row, "Material Name"),
        "materialCode": pick(row, "Material Code"),
        "unit": pick(row, "Unit"),
        "qty": as_number(pick(row, "Qty")),
        "siteId": pick(row, "SiteID"),
        "siteName": pick(row, "Site Name"),
        "doNumber": pick(row, "DO number", "DO Number"),
        "dnNumber": pick(row, "DN Number"),
        "condition": pick(row, "Condition"),
        "picDelivery": pick(row, "PIC Delivery"),
        "vendorSupplier": pick(row, "Vendor Supplier"),
        "idCard": pick(row, "ID Card"),
        "carPlate": pick(row, "Car Plate"),
        "remarks": pick(row, "Remarks"),
        "taggingManual": pick(row, "Tagging Manual"),
        "cableLengthMarker": pick(row, "Cable Length Marker"),
        "cableRoll": pick(row, "Cable Roll"),
        "inOutQty": pick(row, "In-Out Qty"),
        "loCriteria": pick(row, "LO Criteria"),
    }


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, keep_vba=False)
    dv_rows = read_table(wb, "DV", "DV_S")
    material_rows = [map_material(row) for row in read_table(wb, "Sum_WH", "Master_S")]
    do_rows = [map_do(row) for row in read_table(wb, "DO_DB", "DO_DB")]
    site_rows = [map_site(row) for row in read_table(wb, "Site_DB", "Site_DB")]
    log_rows = [map_transaction(row, "logfile") for row in read_table(wb, "Logfile", "Logfile_S")]
    leftover_rows = [map_transaction(row, "leftovers") for row in read_table(wb, "LO_Logfile", "LO_Logfile")]

    master = {
        "transactionTypes": unique([pick(row, "Transaction  Type", "Transaction Type") for row in dv_rows]),
        "taggingTypes": unique([pick(row, "Tagging Type") for row in dv_rows]),
        "units": unique([pick(row, "Unit") for row in dv_rows]),
        "sources": unique([pick(row, "Material Source / Destination") for row in dv_rows]),
        "typeMaterials": unique([pick(row, "Type Material") for row in dv_rows]),
        "conditions": unique([pick(row, "Condition") for row in dv_rows]),
        "warehouses": [
            {
                "whId": pick(row, "WH ID"),
                "whGci": pick(row, "WH GCI"),
                "picWh": pick(row, "PIC WH"),
            }
            for row in dv_rows
            if pick(row, "WH GCI")
        ],
        "labelPrefixes": [
            {
                "material": pick(row, "Material Leftovers"),
                "prefix": pick(row, "Label ID Prefix"),
            }
            for row in dv_rows
            if pick(row, "Material Leftovers") and pick(row, "Label ID Prefix")
        ],
        "cableRolls": unique([pick(row, "Cable Roll") for row in dv_rows]),
        "materialMilestones": unique([pick(row, "Material Milestone") for row in dv_rows]),
    }

    payload = {
        "sourceWorkbook": WORKBOOK.name,
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "master": master,
        "materials": material_rows,
        "deliveryOrders": do_rows,
        "sites": site_rows,
        "logRows": log_rows,
        "leftoverRows": leftover_rows,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Saved {OUT}")
    print(
        f"materials={len(material_rows)} deliveryOrders={len(do_rows)} "
        f"sites={len(site_rows)} logRows={len(log_rows)} leftoverRows={len(leftover_rows)}"
    )


if __name__ == "__main__":
    main()
